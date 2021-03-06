# -*- coding: utf-8 -*-
from collections import defaultdict
import copy
import os
from decimal import Decimal
from datetime import datetime, date
import re

from django import forms
from django.conf import settings
from django.core.exceptions import ValidationError

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from xlrd import open_workbook, XLRDError

from dashboard.apps.dashboard.constants import COST_TYPES, PAYROLL_COSTS
from dashboard.apps.dashboard.spreadsheets import Export, CURRENCY_FORMAT
from dashboard.libs.date_tools import get_workdays
from dashboard.apps.dashboard.models import Person, Rate, Product, PersonCost

from .widgets import DateRangeWidget, RE_DATE_RANGE


REQUIRED_COLS = [
    'Init',
    'Surname',
    'Salary',
    'Staff',
    'Write Offs',
]

BASE_SALARY_RATE_KEY = 'Salary'


def year_range(backward=0, forward=10):
    this_year = date.today().year
    return range(this_year-backward, this_year+forward)


class DateRangeField(forms.DateField):
    input_formats = ['%Y-%m-%d']
    widget = DateRangeWidget

    def format(self, v):
        return datetime.strptime(v, self.input_formats[0]).date()

    def to_python(self, value):
        """
        Validates that the input can be converted to a date. Returns a Python
        tuple of two datetime.date objects.
        """
        if value:
            if RE_DATE_RANGE.match(value):
                value = value.split(':')
                f, t = (self.format(v) for v in value)
                if f < t:
                    return f, t
                raise ValidationError('From date must be before To date.')

            raise ValidationError('Must be two dates')

    def validate(self, value):
        def validate_date(v):
            if isinstance(v, date):
                return v
            try:
                return self.format(v)
            except (TypeError, ValueError):
                raise ValidationError('Invalid date.')

        try:
            from_date, to_date = map(validate_date, value)
        except (TypeError, ValueError):
            raise ValidationError('Date range must have two dates')

        return from_date, to_date


class PayrollUploadForm(forms.Form):
    date_range = DateRangeField(required=True)
    payroll_file = forms.FileField(
        required=True,
        help_text='This must be an original payroll file from HR in .xls '
                  'format')

    @property
    def month(self):
        dr = self.cleaned_data.get('date_range', None)
        if dr:
            return dr[0].strftime('%Y-%m')

    def get_person(self, row, data):
        try:
            return Person.objects.get(
                staff_number=data['Staff'],
            )
        except Person.DoesNotExist:
            pass

        try:
            return Person.objects.get(
                name__icontains=data.get('Surname'),
                name__startswith=data.get('Init')[0],
                is_contractor=False,
            )
        except Person.DoesNotExist:
            try:
                return Person.objects.get(
                    name__icontains=data.get('Surname'),
                    is_contractor=False,
                )
            except (Person.DoesNotExist, Person.MultipleObjectsReturned):
                pass
            self.add_error(
                'payroll_file',
                'ERROR ROW %s: Civil Servant not found with Surname "%s" '
                'and initials "%s"' %
                (row, data.get('Surname'), data.get('Init')))
        except Person.MultipleObjectsReturned:
            self.add_error(
                'payroll_file',
                'ERROR ROW %s: Multiple Civil Servants found with Surname '
                '"%s"' % (row, data.get('Surname')))

    def clean_payroll_file(self):
        start, end = self.cleaned_data.get('date_range')
        if not start or not end:
            raise ValidationError('No date supplied.')

        try:
            workbook = open_workbook(
                file_contents=self.cleaned_data['payroll_file'].read())
        except XLRDError:
            raise ValidationError('Could not read .xls file.')

        ws = workbook.sheet_by_index(0)

        headers = ws.row_values(1)

        missing_required = []
        for required in PAYROLL_COSTS + REQUIRED_COLS:
            if required not in headers:
                missing_required.append(required)

        if missing_required:
            raise ValidationError('Missing fields in upload: %s' % ', '
                                  .join(missing_required))

        payroll = []

        for row in range(2, ws.nrows):
            row_data = ws.row_values(row)
            if all([str(v) for v in row_data[:len(headers) - 1]]):
                data = dict(zip(headers, row_data))
                person = self.get_person(row, data)
                if person:
                    day_rate = Decimal(data[BASE_SALARY_RATE_KEY]) / get_workdays(
                        start, end)
                    staff_number = int(data['Staff'])

                    additional = {}
                    for header in PAYROLL_COSTS:
                        if data[header] and Decimal(data[header]) != 0:
                            additional[header] = Decimal(data[header])

                    if data['Write Offs'] and Decimal(data['Write Offs']) != 0:
                        additional['Write Offs'] = -Decimal(data['Write Offs'])

                    payroll.append({
                        'person': person,
                        'rate': day_rate,
                        'start': start,
                        'end': end,
                        'staff_number': staff_number,
                        'additional': additional
                    })

        return payroll

    def save(self):
        for pay in self.cleaned_data['payroll_file']:
            person = pay['person']
            rate, created = Rate.objects.get_or_create(
                person=person,
                start_date=pay['start'],
                defaults=dict(rate=pay['rate'],)
            )

            if not created:
                rate.rate = pay['rate']
                rate.save()

            if not person.staff_number:
                person.staff_number = pay['staff_number']
                person.save()

            for name, cost in pay['additional'].items():
                additional, created = PersonCost.objects.get_or_create(
                    person=person,
                    start_date=pay['start'],
                    end_date=pay['end'],
                    type=COST_TYPES.MONTHLY,
                    name=name,
                    defaults=dict(cost=cost),
                )
                additional.cost = cost
                additional.save()


EXPORTS = (
    ('Adjustment_Journal', 'Adjustment Export'),
    ('Intercompany_Journal', 'Intercompany Export'),
    ('Product_Detail', 'Product Detail Export'),
)


class ExportForm(forms.Form):
    date_range = DateRangeField(required=True)
    product = forms.ModelChoiceField(
        queryset=Product.objects.all().order_by('area', 'name'),
        required=True)
    export_type = forms.ChoiceField(choices=EXPORTS, widget=forms.RadioSelect())

    def export(self):
        ExportClass = EXPORT_CLASSES.get(self.cleaned_data['export_type'])
        export = ExportClass(cleaned_data=self.cleaned_data)
        return export.export()


class TemplateExport(Export):
    template = 'xls/Journal_Template.xltm'

    def export(self):
        wb = load_workbook(self._get_template(), keep_vba=True)
        self.write(wb)
        return wb

    def _get_template(self):
        for template_setting in settings.TEMPLATES:
            for dir in template_setting['DIRS']:
                template = os.path.join(dir, self.template)
                if os.path.isfile(template):
                    return template
        raise Exception('Could not find template %s' % self.template)  # pragma: no cover

    def write(self, workbook, ws=None):
        if ws is None:  # pragma: no cover
            ws = workbook.get_active_sheet()
        product = self.cleaned_data['product']
        start_date, end_date = self.cleaned_data['date_range']
        month = start_date.strftime('%B \'%y')
        period_start = start_date if start_date.month < 4 else \
            start_date.replace(year=start_date.year + 1)
        period_end = period_start.replace(year=period_start.year + 1)
        period = '%s-%s' % (period_start.strftime('%y'),
                            period_end.strftime('%y'))

        ws.cell(row=11, column=9).value = end_date.strftime('%d/%m/%Y')
        ws.cell(row=12, column=9).value = '%s%s' % (start_date.strftime('%B'),
                                                    period)
        ws.cell(row=13, column=9).value = 'MA100_LW_%s_13' % \
                                          start_date.strftime('%d%m%y')

        def desc(t):
            return '%s - %s Share of DS %s Costs %s' % (
                product.name, product.area.name, t, month)

        # Journal Description
        ws.cell(row=14, column=9).value = desc('')

        # Content Rows
        self.row = 161

        def write_content_row(ct, cost, cc=9):
            ws.cell(row=self.row, column=cc).value = cost
            ws.cell(row=self.row, column=cc).number_format = CURRENCY_FORMAT
            ws.cell(row=self.row, column=10).value = desc(ct)
            ws.cell(row=self.row, column=7).value = product.hr_id
            self.row += 1

        # Total Contractor costs
        write_content_row('Agency', product.people_costs(
            start_date, end_date, contractor_only=True))

        # Total Salary Costs
        write_content_row('Salary', product.non_contractor_salary_costs(
            start_date, end_date))

        # Other Costs
        for cost_type in PAYROLL_COSTS:
            write_content_row(cost_type, product.people_additional_costs(
                start_date, end_date, name=cost_type))

        # Total Write Offs
        write_content_row('Write Offs', -product.people_additional_costs(
            start_date, end_date, name='Write Offs'), 8)

        # Total People costs for project
        write_content_row('Resource', product.people_costs(
            start_date, end_date), 8)


class AdjustmentExport(TemplateExport):
    def write(self, workbook, ws=None):
        ws = workbook.get_active_sheet()
        super(AdjustmentExport, self).write(workbook, ws=ws)
        ws.cell(row=8, column=9).value = 'Adjustment'


class IntercompanyExport(TemplateExport):
    def write(self, workbook, ws=None):
        ws = workbook.get_active_sheet()
        super(IntercompanyExport, self).write(workbook, ws=ws)
        ws.cell(row=8, column=9).value = 'Intercompany Transfer'


class ProductDetailExport(TemplateExport):
    template = 'xls/ProductDetail.xlsx'

    def write(self, workbook, ws=None):
        ws = workbook.get_active_sheet()
        product = self.cleaned_data['product']
        start_date, end_date = self.cleaned_data['date_range']

        details = defaultdict(lambda: defaultdict(Decimal))

        def add_cost(task):
            s = max(task.start_date, start_date)
            e = min(task.end_date, end_date)
            if not task.person.is_contractor:
                for name in PAYROLL_COSTS:
                    details[task.person][name] += task.people_costs(
                        s, e, additional_cost_name=name)
            details[task.person]['total'] += task.people_costs(s, e)
            details[task.person]['days'] += task.get_days(s, e)
            details[task.person][BASE_SALARY_RATE_KEY] = \
                task.person.base_rate_between(s, e) * \
                details[task.person]['days']

        list(map(add_cost, product.tasks.between(start_date, end_date)))

        ws.cell(row=2, column=1).value = datetime.now().strftime('%d/%m/%Y %I:%M%p')
        ws.cell(row=3, column=7).value = '%s DRAFT' % \
                                         start_date.strftime('%b %Y').upper()
        ws.cell(row=4, column=1).value = 'Product: %s' % product.name

        initial_row = 8
        row = initial_row
        for person, detail in details.items():
            ws.cell(row=row, column=1).value = person.name
            ws.cell(row=row, column=2).value = person.job_title
            ws.cell(row=row, column=3).value = product.name
            ws.cell(row=row, column=4).value = product.area.name

            ws.cell(row=row, column=10).value = detail['days'] * Decimal('8')
            ws.cell(row=row, column=11).value = detail['days']

            ws.cell(row=row, column=12).value = ''

            if person.is_contractor:
                ws.cell(row=row, column=13).value = person.rate_between(
                    start_date, end_date)
            ws.cell(row=row, column=14).value = detail['total']

            ws.cell(row=row, column=15).value = ''

            if not person.is_contractor:
                ws.cell(row=row, column=18).value = detail['Salary']
                ws.cell(row=row, column=19).value = detail['Misc.Allow.']
                ws.cell(row=row, column=20).value = detail['ERNIC']
                ws.cell(row=row, column=21).value = '?'

            insert_rows(ws, row, 1)
            row += 1

        for col in ['J', 'K', 'N']:
            cells = []
            for r in range(initial_row, row):
                cells.append('%s%s' % (col, r))

            self.write_sum(ws, cells, '%s%s' % (col, row + 1))
            if col == 'N':
                self.write_sum(ws, cells, '%s%s' % (col, row + 6))

    def write_sum(self, ws, cells, coordinate):
        cell = ws.cell(coordinate=coordinate)
        cell.set_explicit_value(
            value='=SUM(%s)' % ','.join(cells), data_type=cell.TYPE_FORMULA)


EXPORT_CLASSES = {
    'Adjustment_Journal': AdjustmentExport,
    'Intercompany_Journal': IntercompanyExport,
    'Product_Detail': ProductDetailExport,
}


def insert_rows(ws, row_idx, cnt, above=False, copy_style=True,
                fill_formulae=True):  # flake8: noqa  # pragma: no cover
    """Inserts new (empty) rows into worksheet at specified row index.

    :param row_idx: Row index specifying where to insert new rows.
    :param cnt: Number of rows to insert.
    :param above: Set True to insert rows above specified row index.
    :param copy_style: Set True if new rows should copy style of immediately
    above row.
    :param fill_formulae: Set True if new rows should take on formula from
    immediately above row, filled with references new to rows.

    Usage:

    * insert_rows(2, 10, above=True, copy_style=False)

    """
    RE_RANGE = re.compile(
        "(?P<s_col>[A-Z]+)(?P<s_row>\d+):(?P<e_col>[A-Z]+)(?P<e_row>\d+)")

    row_idx = row_idx - 1 if above else row_idx

    old_cells = set()
    old_fas = set()
    new_cells = dict()
    new_fas = dict()
    for c in ws._cells.values():
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row, c.col_idx))
            c.row += cnt
            new_cells[(c.row, c.col_idx)] = c
            if c.data_type == Cell.TYPE_FORMULA:
                c.value = re.sub(
                    "(\$?[A-Z]{1,3})\$?%d" % (c.row - cnt),
                    lambda m: m.group(1) + str(c.row),
                    c.value
                )
                if old_coor in ws.formula_attributes:
                    old_fas.add(old_coor)
                    fa = ws.formula_attributes[old_coor].copy()
                    if 'ref' in fa:
                        if fa['ref'] == old_coor:
                            fa['ref'] = c.coordinate
                        else:
                            g = RE_RANGE.search(fa['ref']).groupdict()
                            fa['ref'] = g['s_col'] + str(
                                int(g['s_row']) + cnt) + ":" + g[
                                            'e_col'] + str(
                                int(g['e_row']) + cnt)
                    new_fas[c.coordinate] = fa

    for coor in old_cells:
        del ws._cells[coor]
    ws._cells.update(new_cells)

    for fa in old_fas:
        del ws.formula_attributes[fa]
    ws.formula_attributes.update(new_fas)

    for row in range(len(ws.row_dimensions) - 1 + cnt, row_idx + cnt, -1):
        new_rd = copy.copy(ws.row_dimensions[row - cnt])
        new_rd.index = row
        ws.row_dimensions[row] = new_rd
        del ws.row_dimensions[row - cnt]

    row_idx += 1
    for row in range(row_idx, row_idx + cnt):
        new_rd = copy.copy(ws.row_dimensions[row - 1])
        new_rd.index = row
        ws.row_dimensions[row] = new_rd
        for col in range(1, ws.max_column):
            col = get_column_letter(col)
            cell = ws.cell('%s%d' % (col, row))
            cell.value = None
            source = ws.cell('%s%d' % (col, row - 1))
            if copy_style:
                cell.number_format = source.number_format
                cell.font = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border = source.border.copy()
                cell.fill = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                if source.value == "=":
                    if source.coordinate in ws.formula_attributes:
                        fa = ws.formula_attributes[source.coordinate].copy()
                        ws.formula_attributes[cell.coordinate] = fa
                else:
                    cell.value = re.sub(
                        "(\$?[A-Z]{1,3})\$?%d" % (row - 1),
                        lambda m: m.group(1) + str(row),
                        source.value
                    )
                cell.data_type = Cell.TYPE_FORMULA

    for cr_idx, cr in enumerate(ws.merged_cell_ranges):
        g = RE_RANGE.search(cr).groupdict()
        if row_idx >= int(g['s_row']) and row_idx <= int(g['e_row']):
            ws.merged_cell_ranges[cr_idx] = g['s_col'] + g['s_row'] + ":" + g[
                'e_col'] + str(int(g['e_row']) + cnt)

    for k, v in ws.formula_attributes.items():
        if 'ref' in v:
            ref = v['ref']
            if ":" in ref:
                g = RE_RANGE.search(v['ref']).groupdict()
                if row_idx >= int(g['s_row']) and row_idx <= int(g['e_row']):
                    ws.formula_attributes[k]['ref'] = g['s_col'] + g[
                        's_row'] + ":" + g['e_col'] + str(
                        int(g['e_row']) + cnt)
